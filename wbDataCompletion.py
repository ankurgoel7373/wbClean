import openpyxl
import glob
import datetime
from dateutil.parser import parse
import sys
import os
import time
import operator
import signal
from collections import defaultdict

ErrorDictionary = {'AVB001': 'ALL COLUMNS NONE','AVB002': 'Error Parsing Row','AVB003': 'Error Parsing Date'}


path = "N:\\Clean West Bengal\\Individual Files\\*.xlsx"

dictionaryWestBengal = "M:\\Clean West Bengal\\Data Dictionary WestBengal.xlsx"
masterWestBengal="M:\\Clean West Bengal\\MasterWestBengal.xlsx"

CacheMasterWestBengal = "M:\\Clean West Bengal\\CacheMasterWestBengal.xlsx"

TPAMaster = {12:'Medicare Insurance TPA Services (India) Pvt. Ltd',20:'Genins India Insurance TPA Limited',26:'SAFEWAY INSURANCE TPA (P) LTD',8:'Heritage Health Insurance TPA Pvt Ltd',16:'Vidal Health Insurance Third Party Administrator TPA Pvt. Ltd.',28:'Dedicated Healthcare Services TPA (I) Pvt. Ltd',19:'MedSave Health Insurance TPA Limited'}

CodeNameColumnDict = {5:4, 3:2}
ColumnMasterMapping= {3:'Insurance Company Master',5:'TPA Master'}

CacheMasterDictionary = {}

#MasterNameColumn = [4]
#MasterCodeColumn=[5]

StandardColumns = ['S No','Insurance Coompany Name','Insurance Coompany Code','TPA Name','TPA Code','Policy No','Risk Start Date','Patient ID card Number','Member Name / Patient Name','Family Head ID card Number',' Family Head Name','Mobile No','Gender','Age','Relationship','Address','State','District Name','Pin Code','URN','Village Name','Block Name','Panchayat Name','Aadhaar No','Bill No','IP No','TransactionDate','TransactionTime','Transaction Code','PackageCode','PackageName','PackageCost','No Of Days','Upload Dt','Mobile No Claiment','Claim ID','Date of Addmission System Date & Time','Date of Discharge System Date & Time','Total Amount Blocked','Amount Claimed','Disallowed  Amount','Settled Amount','TDS Rate','TDS Amount','Final paid Amount','float No','Float Request  date','Payment date','NEFT / DD No',' Payee Name','Discharge Diagnosis','Procedure code','Procedure Type','Surgeon Name','Doctor Registration No','Doctor Mobile No','Patient Registration No','Surgery Date','Surgery Time','Provider Type','Provider Grade','Provider ID','Provider Name','Provider adress','Provider State','Provider District','Provider Pin Code','Provider PAN','Hospital Mobile No','Disallowance Reason','Pending Reason','Rejection Reason','Rejection Date','Query Raised Date','Query Rcvd Date','Reopen Date','Claim Investigation (Y / N) ','Claim Investigation date','Claim Investigation Report Summary','PreAuth Status','Claim Processing_Status','Remarks','Call to Beneficiary (Y/N)','Date of Call','Remarks of Call']
whitelist = set('abcdefghijklmnopqrstuvwxy ABCDEFGHIJKLMNOPQRSTUVWXYZ')


class GracefulKiller:
  kill_now = False
  def __init__(self):
    signal.signal(signal.SIGINT, self.exit_gracefully)
    signal.signal(signal.SIGTERM, self.exit_gracefully)

  def exit_gracefully(self,signum, frame):
    self.kill_now = True

def initCacheMasterWestBengal():
    wb = openpyxl.load_workbook(CacheMasterWestBengal,data_only=True)
    for sheet,name in zip(wb.worksheets,wb.get_sheet_names()):
        tempDictionary = {}
        for row in sheet.iter_rows(row_offset=1):
            if any(cell.value for cell in row):
                tempDictionary[row[0].value]=row[1].value
        CacheMasterDictionary[name] = tempDictionary

def initMasterWestBengal():

    masterDictionary={}
    wb = openpyxl.load_workbook(masterWestBengal,data_only=True)
    for sheet,name in zip(wb.worksheets,wb.get_sheet_names()):
        tempDictionary = {}
        for row in sheet.iter_rows(row_offset=1):
            if any(cell.value for cell in row):
                tempDictionary[row[0].value]=row[1].value
        masterDictionary[name]=tempDictionary
    return masterDictionary



def maximum(y):
    mymax = y[0][1]
    return [a for a in y if a[1]==mymax]


# here X is given value and Y is the value from the master.

def EqualityFactor(X, Y):
    # find the length of the strings
    m = len(X)
    n = len(Y)

    # declaring the array for storing the dp values
    L = [[None] * (n + 1) for i in range(m + 1)]

    """Following steps build L[m+1][n+1] in bottom up fashion
    Note: L[i][j] contains length of LCS of X[0..i-1]
    and Y[0..j-1]"""
    for i in range(m + 1):
        for j in range(n + 1):
            if i == 0 or j == 0:
                L[i][j] = 0
                #print(L[i][j]," ", end='')
            elif X[i - 1] == Y[j - 1]:
                L[i][j] = L[i - 1][j - 1] + 1
                #print(L[i][j]," ", end='')
            else:
                L[i][j] = max(L[i - 1][j], L[i][j - 1])
                #print(L[i][j]," ", end='')


    Z=[]
    #Compute Equality Factor for every cell of last row.

    for k in range(n,0,-1):
        j=k
        i=m
        x=0
        y=0
        flag=True

        while True:

            if L[i][j]==0:
                break

            if X[i-1]==Y[j-1]:
                if flag:
                   flag=False
                   x=i
                   y=j
                j=j-1
                i=i-1
            else:
                if L[i][j-1]>L[i-1][j]:
                    j=j-1
                else:
                    i=i-1
        Z.append(L[m][k]-abs((x-i)-(y-j)))
    # end of function lcs
    return max(Z)



# master is a dictionary, with code and name values mapped

def processMasterColumn(codeColumn, nameColumn, master,cacheMaster):
    try :
        results=[]
        blanksIdentified=[0,0]
        blanksResolved=[0,0]
        ValueOutOfMaster=[0,0]
        ValueResolvedToMaster=[0,0]

        unresolvedCodeDictionary={}
        unresolvedNameDictionary={}

        if codeColumn is None and nameColumn is None:
           blanksIdentified[0]+=1
           blanksIdentified[1]+= 1
           codeColumn=""
           nameColumn=""

        elif codeColumn is None or nameColumn is None:

            if codeColumn is None:
                blanksIdentified[0] += 1

                if nameColumn in master.values():
                    for key in master:
                        if nameColumn==master[key]:
                            codeColumn=key
                            blanksResolved[0]+=1
                            break;
                else:
                    ValueOutOfMaster[1] += 1
                    unresolvedNameDictionary[nameColumn] = unresolvedNameDictionary.get(nameColumn, 0) + 1

                    equalityFactorArray = []
                    for key in master:
                        Z=EqualityFactor(nameColumn.lower(),master[key].lower())
                        equalityFactorArray.append((key,Z))
                    equalityFactorArray.sort(key=operator.itemgetter(1), reverse=True)
                    equalityFactorArray=maximum(equalityFactorArray)
                    if(len(equalityFactorArray)==1):
                        codeColumn=equalityFactorArray[0][0]
                        cacheMaster[nameColumn]=codeColumn
                        nameColumn=master[codeColumn]
                        blanksResolved[0]+=1
                        ValueResolvedToMaster[1]+1
            else:
                blanksIdentified[1] += 1

                if codeColumn in master:
                   nameColumn = master[codeColumn]
                   blanksResolved[1]+=1
                else:
                    nameColumn=""
                    ValueOutOfMaster[0]+=1
                    unresolvedCodeDictionary[codeColumn] = unresolvedCodeDictionary.get(codeColumn,0)+1

        else:

            if not (codeColumn in master and nameColumn in master.values()):

                if codeColumn in master:
                    ValueOutOfMaster[1]+=1
                    unresolvedNameDictionary[nameColumn] = unresolvedNameDictionary.get(nameColumn, 0) + 1
                    nameColumn = master[codeColumn]
                    ValueResolvedToMaster[1] += 1

                elif nameColumn in master.values():

                    ValueOutOfMaster[0] += 1
                    unresolvedCodeDictionary[codeColumn] = unresolvedCodeDictionary.get(codeColumn, 0) + 1

                    for key in master:
                        if nameColumn == master[key]:
                            codeColumn = key
                            ValueResolvedToMaster[0]+=1
                            break;
                else:
                    ValueOutOfMaster[0] += 1
                    ValueOutOfMaster[1] += 1
                    unresolvedCodeDictionary[codeColumn] =unresolvedCodeDictionary.get(codeColumn, 0) + 1
                    unresolvedNameDictionary[nameColumn] = unresolvedNameDictionary.get(nameColumn, 0) + 1

                    equalityFactorArray=[]

                    for key in master:
                        Z = EqualityFactor(nameColumn.lower(), master[key].lower())
                        equalityFactorArray.append((key, Z))
                    equalityFactorArray.sort(key=operator.itemgetter(1), reverse=True)
                    equalityFactorArray = maximum(equalityFactorArray)

                    if (len(equalityFactorArray) == 1):
                        codeColumn = equalityFactorArray[0][0]
                        cacheMaster[nameColumn] = codeColumn
                        nameColumn=master[codeColumn]
                        ValueResolvedToMaster[0] += 1
                        ValueResolvedToMaster[1] += 1

        if nameColumn is None:
            codeColumn = ""
        if nameColumn is None:
            nameColumn = ""

        results.append(codeColumn)
        results.append(nameColumn)
        return results
    except Exception as e:
        print(e)


def cell_value(string, cell_number):
    if string is None:
        return ""
    else:
        string=str(string)

    policyStartDate = datetime.datetime.strptime("2017-02-01", '%Y-%m-%d').strftime('%Y-%m-%d')
    possibleFormats = ['%Y-%m-%d%H:%M:%S','%d/%m/%Y%H:%M:%S', '%d/%m/%Y', '%Y-%m-%d','%d%b%Y', '%d%B%Y','%d-%b-%Y', '%d-%B-%Y']

    dateColumns = [27, 34, 37, 38, 47, 48]
    amountColumns = [39, 40, 41, 42, 43, 44, 45]

    if cell_number in dateColumns:
        string = string.replace("00:00:00", "")
        string = string.replace(" ", "")
        ans = string
        flag = 0
        for format in possibleFormats:
            try:
                if flag == 0:
                    ans = datetime.datetime.strptime(string, format).strftime('%Y-%m-%d')
                    if ans < policyStartDate:
                        print("Orignal Date:"+string+", Parsed Date:"+ans)
                        print("Error, Date is less than policy start date")
                    flag = 1

            except:
                continue

        return ans

    elif cell_number in amountColumns:
        try:
            amount=float(string)
        except:
            amount=string
        return amount
    else:
        return string


def header_row(row, invalidColumn):
    MatchCount = 0
    columnIndex = 1
    stdColIndex = 0
    try:
        for cell in row:
            if columnIndex in invalidColumn:
                columnIndex += 1
                continue
            else:
                temp1=StandardColumns[stdColIndex].lower().replace(" ","")
                temp2=(str(cell.value)).lower().replace(" ","")
                if temp1 in temp2 or temp2 in temp1:
                    #print(temp1+ " and "+ temp2)
                    MatchCount += 1
                stdColIndex += 1
                columnIndex += 1
    except:
        print(MatchCount)
    if MatchCount > 45:
        #print("True")
        return True
    else:
        #print("False")
        return False


def convertNumeric(str):
    if type(str) is int:
        return str
    else:
        try:
            ans = int(str)
            return ans
        except:
            return str

def main(files, saveFileName, processedTime,MasterDictionary):

    os.mkdir("N:\\Clean West Bengal\\Processed Files\\"+processedTime)

    MergeFile = openpyxl.Workbook()
    MergeSheet=MergeFile.active
    rowMerge = 1;
    columnMerge = 1;

    for col in range(len(StandardColumns)):
        MergeSheet.cell(row=rowMerge,column=col+1).value=StandardColumns[col]
    rowMerge += 1

    for file in files:
        print("Parsing File : " + file)

        wb = openpyxl.load_workbook(file, data_only=True)
        for sheet,name in zip(wb.worksheets,wb.get_sheet_names()):
            if "format" not in name.lower():

                print("Parsing sheet : " + name)

                columnCount = sheet.max_column
                if columnCount == 85:
                    for row in sheet.iter_rows(row_offset=1):
                        try:
                            if not any(cell.value for cell in row):
                                print("Data Missing")
                                continue
                            if header_row(row,[0]):
                                print("Repeated Data Header Row")
                                continue
                            columnMerge = 1;

                            for idx, cell in enumerate(row):
                                if columnMerge in CodeNameColumnDict.values():
                                    columnMerge += 1
                                    continue
                                if columnMerge in CodeNameColumnDict:
                                   nameColumn = row[CodeNameColumnDict[columnMerge]-1].value
                                   codeColumn = cell.value
                                   resultColumns = processMasterColumn(convertNumeric(codeColumn),nameColumn,MasterDictionary[ColumnMasterMapping[columnMerge]],CacheMasterDictionary[ColumnMasterMapping[columnMerge]])
                                   MergeSheet.cell(row=rowMerge, column=columnMerge).value = resultColumns[0]
                                   MergeSheet.cell(row=rowMerge, column=CodeNameColumnDict[columnMerge]).value = resultColumns[1]
                                   columnMerge += 1
                                else:
                                    MergeSheet.cell(row=rowMerge,column=columnMerge).value = cell_value(cell.value,columnMerge)
                                    columnMerge += 1
                            rowMerge += 1
                        except Exception as e:
                            print("Error parsing row in sheet:" + name)
                            print(e)
                elif columnCount == 86 and "data" in str(sheet.cell(row=1,column=1).value).lower():
                    try:
                        for row in sheet.iter_rows(row_offset=1):
                            if not any(cell.value for cell in row):
                                print("Data Missing")
                                continue
                            if header_row(row, [1]):
                                print("Repeated Data Header Row")
                                continue
                            columnMerge = 1;
                            leaveColumn = 0;
                            for idx, cell in enumerate(row):
                                if leaveColumn == 0:
                                    leaveColumn = 1
                                    continue
                                if columnMerge in CodeNameColumnDict.values():
                                    columnMerge += 1
                                    continue
                                if columnMerge in CodeNameColumnDict:
                                   nameColumn = row[CodeNameColumnDict[columnMerge]].value
                                   codeColumn = cell.value
                                   resultColumns = processMasterColumn(convertNumeric(codeColumn),nameColumn,MasterDictionary[ColumnMasterMapping[columnMerge]],CacheMasterDictionary[ColumnMasterMapping[columnMerge]])
                                   MergeSheet.cell(row=rowMerge, column=columnMerge).value = resultColumns[0]
                                   MergeSheet.cell(row=rowMerge, column=CodeNameColumnDict[columnMerge]).value = resultColumns[1]
                                   columnMerge += 1
                                else:
                                    MergeSheet.cell(row=rowMerge,column=columnMerge).value = cell_value(cell.value,columnMerge)
                                    columnMerge += 1
                            rowMerge += 1
                    except:
                        print("Error parsing row in sheet:"+name)
                else:
                    invalidColumn = []
                    stdCol = 0
                    for col in range(sheet.max_column):

                        if not((''.join(filter(whitelist.__contains__,str(sheet.cell(row=1,column=col+1).value))).lower().replace(" ","")) == (''.join(filter(whitelist.__contains__, StandardColumns[stdCol])).lower().replace(" ",""))):
                            invalidColumn.append(col+1)
                        else:
                            stdCol += 1
                    print("Invalid Columns in FileName, SheetName:"+file+","+name)
                    print(invalidColumn)
                    if(sheet.max_column-len(invalidColumn) == 85):
                        try:
                            for row in sheet.iter_rows(row_offset=1):
                                if not any(cell.value for cell in row):
                                    print("Data Missing")
                                    continue
                                if header_row(row, invalidColumn):
                                    print("Repeated Data Header Row")
                                    continue
                                columnMerge = 1;
                                leaveColumn = 1;

                                for idx, cell in enumerate(row):
                                    if leaveColumn in invalidColumn:
                                        leaveColumn += 1
                                        continue
                                    if columnMerge in CodeNameColumnDict.values():
                                        columnMerge += 1
                                        leaveColumn += 1
                                        continue
                                    if columnMerge in CodeNameColumnDict:
                                        nameColumn = row[CodeNameColumnDict[columnMerge]+(leaveColumn-columnMerge)-1].value
                                        codeColumn = cell.value
                                        resultColumns = processMasterColumn(convertNumeric(codeColumn), nameColumn,MasterDictionary[ColumnMasterMapping[columnMerge]],CacheMasterDictionary[ColumnMasterMapping[columnMerge]])
                                        MergeSheet.cell(row=rowMerge, column=columnMerge).value = resultColumns[0]
                                        MergeSheet.cell(row=rowMerge, column=CodeNameColumnDict[columnMerge]).value =resultColumns[1]
                                        columnMerge += 1
                                    else:
                                        MergeSheet.cell(row=rowMerge, column=columnMerge).value = cell_value(cell.value,columnMerge)
                                        columnMerge += 1
                                    leaveColumn += 1
                                rowMerge += 1
                        except:
                            print("Error parsing row in sheet:"+name)
                    else:
                        print(file+":"+name+": Some Columns may be missing or column naming conventions may not be followed")

            else:
                print("Sheet Not Merged:"+file+":" + name)
        os.rename(file, file.replace("Individual Files","Processed Files\\"+processedTime))

    MergeFile.save("N:\\Clean West Bengal\\Merged Files\\"+saveFileName)
    print("File Saved Successfully")
    print(CacheMasterDictionary)

def FinalMain():
    MasterDictionary = initMasterWestBengal()
    initCacheMasterWestBengal()
    killer = GracefulKiller()
    while True:
        files = glob.glob(path)
        if killer.kill_now:
            break
        elif len(files) >0:
            present = datetime.datetime.now()
            main(files,"MergeFile"+present.strftime('%d_%m_%Y_%H_%M_%S')+".xlsx",present.strftime('%d_%m_%Y_%H_%M_%S'),MasterDictionary)
        else:
            time.sleep(60)

    print("Program exiting Gracefully")
    wb = openpyxl.load_workbook(CacheMasterWestBengal, data_only=True)
    for CMD,sheet, name in zip(CacheMasterDictionary,wb.worksheets, wb.get_sheet_names()):
        rowCacheSheet = 2
        for key in CMD:
            sheet.cell(row=rowCacheSheet,column=1).value=key
            sheet.cell(row=rowCacheSheet,column=2).value=CMD[key]
            rowCacheSheet += 1
    wb.save(CacheMasterWestBengal)

FinalMain()