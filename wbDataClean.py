import openpyxl
import glob
import datetime
from dateutil.parser import parse
import sys
import os
import time

ErrorDictionary = {'AVB001': 'ALL COLUMNS NONE','AVB002': 'Error Parsing Row','AVB003': 'Error Parsing Date'}



path = "N:\\Clean West Bengal\\Individual Files\\*.xlsx"


StandardColumns = ['S No','Insurance Coompany Name','Insurance Coompany Code','TPA Name','TPA Code','Policy No','Risk Start Date','Patient ID card Number','Member Name / Patient Name','Family Head ID card Number',' Family Head Name','Mobile No','Gender','Age','Relationship','Address','State','District Name','Pin Code','URN','Village Name','Block Name','Panchayat Name','Aadhaar No','Bill No','IP No','TransactionDate','TransactionTime','Transaction Code','PackageCode','PackageName','PackageCost','No Of Days','Upload Dt','Mobile No Claiment','Claim ID','Date of Addmission System Date & Time','Date of Discharge System Date & Time','Total Amount Blocked','Amount Claimed','Disallowed  Amount','Settled Amount','TDS Rate','TDS Amount','Final paid Amount','float No','Float Request  date','Payment date','NEFT / DD No',' Payee Name','Discharge Diagnosis','Procedure code','Procedure Type','Surgeon Name','Doctor Registration No','Doctor Mobile No','Patient Registration No','Surgery Date','Surgery Time','Provider Type','Provider Grade','Provider ID','Provider Name','Provider adress','Provider State','Provider District','Provider Pin Code','Provider PAN','Hospital Mobile No','Disallowance Reason','Pending Reason','Rejection Reason','Rejection Date','Query Raised Date','Query Rcvd Date','Reopen Date','Claim Investigation (Y / N) ','Claim Investigation date','Claim Investigation Report Summary','PreAuth Status','Claim Processing_Status','Remarks','Call to Beneficiary (Y/N)','Date of Call','Remarks of Call']
whitelist = set('abcdefghijklmnopqrstuvwxy ABCDEFGHIJKLMNOPQRSTUVWXYZ')


def cell_value(string, cell_number):
    if string is None:
        return ""
    else:
        string=str(string)

    policyStartDate = datetime.datetime.strptime("2017-02-01", '%Y-%m-%d').strftime('%Y-%m-%d')
    possibleFormats = ['%Y-%m-%d%H:%M:%S','%d/%m/%Y%H:%M:%S', '%d/%m/%Y', '%Y-%m-%d','%d%b%Y', '%d%B%Y','%d-%b-%Y', '%d-%B-%Y']
    dateColumns = [27, 34, 37, 38, 47, 48]
    amountColumns = [39, 40, 41, 42, 43, 44, 45]

    if cell_number in dateColumns:
        string = string.replace("00:00:00", "")
        string = string.replace(" ", "")
        ans = string
        flag = 0
        for format in possibleFormats:
            try:
                if flag == 0:
                    ans = datetime.datetime.strptime(string, format).strftime('%Y-%m-%d')
                    if ans < policyStartDate:
                        print("Orignal Date:"+string+", Parsed Date:"+ans)
                        print("Error, Date is less than policy start date")
                    flag = 1

            except:
                continue

        return ans

    elif cell_number in amountColumns:
        try:
            amount=float(string)
        except:
            amount=string
        return amount
    else:
        return string


def header_row(row, invalidColumn):
    MatchCount = 0
    columnIndex = 1
    stdColIndex = 0
    try:
        for cell in row:
            if columnIndex in invalidColumn:
                columnIndex += 1
                continue
            else:
                temp1=StandardColumns[stdColIndex].lower().replace(" ","")
                temp2=(str(cell.value)).lower().replace(" ","")
                if temp1 in temp2 or temp2 in temp1:
                    #print(temp1+ " and "+ temp2)
                    MatchCount += 1
                stdColIndex += 1
                columnIndex += 1
    except:
        print(MatchCount)
    if MatchCount > 45:
        #print("True")
        return True
    else:
        #print("False")
        return False


def main(files, saveFileName, processedTime):

    os.mkdir("N:\\Clean West Bengal\\Processed Files\\"+processedTime)

    MergeFile = openpyxl.Workbook()
    MergeSheet=MergeFile.active
    rowMerge = 1;
    columnMerge = 1;

    for col in range(len(StandardColumns)):
        MergeSheet.cell(row=rowMerge,column=col+1).value=StandardColumns[col]
    rowMerge += 1

    for file in files:
        print("Parsing File : " + file)

        wb = openpyxl.load_workbook(file, data_only=True)
        for sheet,name in zip(wb.worksheets,wb.get_sheet_names()):
            if "format" not in name.lower():

                print("Parsing sheet : " + name)

                columnCount = sheet.max_column
                if columnCount == 85:
                    for row in sheet.iter_rows(row_offset=1):
                        try:
                            if not any(cell.value for cell in row):
                                print("Data Missing")
                                continue
                            if header_row(row,[0]):
                                print("Repeated Data Header Row")
                                continue
                            columnMerge = 1;
                            for cell in row:
                                MergeSheet.cell(row=rowMerge,column=columnMerge).value = cell_value(cell.value,columnMerge)
                                columnMerge += 1
                            rowMerge += 1
                        except:
                            print("Error parsing row in sheet:" + name)
                elif columnCount == 86 and "data" in str(sheet.cell(row=1,column=1).value).lower():
                    try:
                        for row in sheet.iter_rows(row_offset=1):
                            if not any(cell.value for cell in row):
                                print("Data Missing")
                                continue
                            if header_row(row, [1]):
                                print("Repeated Data Header Row")
                                continue
                            columnMerge = 1;
                            leaveColumn = 0;
                            for cell in row:
                                if leaveColumn == 0:
                                    leaveColumn = 1
                                    continue

                                MergeSheet.cell(row=rowMerge,column=columnMerge).value = cell_value(cell.value, columnMerge)
                                columnMerge += 1
                            rowMerge += 1
                    except:
                        print("Error parsing row in sheet:"+name)
                else:
                    invalidColumn = []
                    stdCol = 0
                    for col in range(sheet.max_column):

                        if not((''.join(filter(whitelist.__contains__,str(sheet.cell(row=1,column=col+1).value))).lower().replace(" ","")) == (''.join(filter(whitelist.__contains__, StandardColumns[stdCol])).lower().replace(" ",""))):
                            invalidColumn.append(col+1)
                        else:
                            stdCol += 1
                    print("Invalid Columns in FileName, SheetName:"+file+","+name)
                    print(invalidColumn)
                    if(sheet.max_column-len(invalidColumn) == 85):
                        try:
                            for row in sheet.iter_rows(row_offset=1):
                                if not any(cell.value for cell in row):
                                    print("Data Missing")
                                    continue
                                if header_row(row, invalidColumn):
                                    print("Repeated Data Header Row")
                                    continue
                                columnMerge = 1;
                                leaveColumn = 1;
                                for cell in row:
                                    if leaveColumn in invalidColumn:
                                        leaveColumn += 1
                                        continue
                                    MergeSheet.cell(row=rowMerge, column=columnMerge).value = cell_value(cell.value, columnMerge)
                                    columnMerge += 1
                                    leaveColumn += 1
                                rowMerge += 1
                        except:
                            print("Error parsing row in sheet:"+name)
                    else:
                        print(file+":"+name+": Some Columns may be missing or column naming conventions may not be followed")

            else:
                print("Sheet Not Merged:"+file+":" + name)
        os.rename(file, file.replace("Individual Files","Processed Files\\"+processedTime))

    MergeFile.save("N:\\Clean West Bengal\\Merged Files\\"+saveFileName)
    print("File Saved Successfully")

while True:
    files = glob.glob(path)
    if len(files) == 18:
        present = datetime.datetime.now()
        main(files,"MergeFile"+present.strftime('%d_%m_%Y_%H_%M_%S')+".xlsx",present.strftime('%d_%m_%Y_%H_%M_%S'))
    else:
        time.sleep(60)

